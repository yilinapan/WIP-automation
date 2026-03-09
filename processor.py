# -*- coding: utf-8 -*-
"""
WIP 轉貼紙 Excel 處理模組
將 WIP 格式的 Excel 轉換為貼紙格式輸出
"""

import re
from io import BytesIO
from pathlib import Path
from typing import Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU

# 貼紙實體尺寸（cm）
HANGTAG_WIDTH_CM = 9.47
HANGTAG_HEIGHT_CM = 8.96
BAG_WIDTH_CM = 9.47
BAG_HEIGHT_CM = 7.64

# 預設貼紙圖片路徑（專案根目錄，支援 .png / .jpg / .jpeg）
def _default_sticker_path(base_name: str) -> Path:
    """依序嘗試 .png、.jpg、.jpeg，回傳第一個存在的路徑。"""
    parent = Path(__file__).resolve().parent
    for ext in (".png", ".jpg", ".jpeg"):
        p = parent / f"{base_name}{ext}"
        if p.exists():
            return p
    return parent / f"{base_name}.png"  # fallback


# 顏色常數（金色輔色4較淺60%、藍色輔色1較淺80%、橙色輔色2較淺80%/60%）
HEADER_FILL = "E6D99C"      # 金色，輔色4，較淺60%
MPO_FILL = "F9CB9C"        # 橙色，輔色2，較淺60%（A欄有 MPO#）
BLOCK_FILL_1 = "DAE3F3"    # 藍色，輔色1，較淺80%
BLOCK_FILL_2 = "FCE4D6"    # 橙色，輔色2，較淺80%
FONT_NAME = "微軟正黑體"
SIZES = ["S", "M", "L", "XL"]

# 欄寬 A=23, B=33, C=19, F=32；列高 Row 1=25
COLUMN_WIDTHS = [23, 33, 19, 8, 10, 32, 12, 10, 10, 12, 12]
ROW1_HEIGHT = 25


def _clean_color(raw: str) -> str:
    """移除英文、數字、#，只保留純中文"""
    if pd.isna(raw) or not isinstance(raw, str):
        return ""
    return re.sub(r"[a-zA-Z0-9#\s]", "", raw).strip()


def _parse_style(style_val: str) -> tuple[str, str]:
    """解析 STYLE 欄位，回傳 (Style ID, MPO ID)"""
    if pd.isna(style_val) or not isinstance(style_val, str):
        return "", ""
    lines = [ln.strip() for ln in str(style_val).split("\n") if ln.strip()]
    style_id = lines[0] if len(lines) >= 1 else ""
    mpo_id = lines[1] if len(lines) >= 2 else ""
    return style_id, mpo_id


def _is_header_row(row_vals: list, template_first: list) -> bool:
    """判斷該列是否與標題列格式一致（比對前幾欄）"""
    if len(row_vals) < 5 or len(template_first) < 5:
        return False
    for i in range(min(5, len(template_first), len(row_vals))):
        a = str(template_first[i]).strip() if pd.notna(template_first[i]) else ""
        b = str(row_vals[i]).strip() if pd.notna(row_vals[i]) else ""
        if a != b:
            return False
    return True


def _find_header_row_indices(df_raw: pd.DataFrame) -> list:
    """動態搜尋與第一條標題列（Excel Row 3）格式相同的所有標題列索引（0-based）。"""
    if len(df_raw) < 3:
        return []
    template = df_raw.iloc[2].tolist()
    indices = []
    for i in range(len(df_raw)):
        if _is_header_row(df_raw.iloc[i].tolist(), template):
            indices.append(i)
    return indices


def _cols_to_ffill(columns: list) -> list:
    """回傳需要做 ffill 的欄位名稱（合併儲存格欄位）"""
    FFILL_KEYWORDS = [
        "style", "品名", "售價", "價格", "factory", "fobs", "單價",
        "color", "顏色", "sketch", "布料", "組分", "印花",
    ]
    result = []
    for c in columns:
        s = str(c).strip().lower()
        if any(kw in s for kw in FFILL_KEYWORDS):
            result.append(c)
    return result


def _get_blocks_from_sheet(df_raw: pd.DataFrame) -> list:
    """依動態標題列將工作表切成多個資料區塊，只對合併儲存格欄位做 ffill。"""
    header_indices = _find_header_row_indices(df_raw)
    if not header_indices:
        header_indices = [2] if len(df_raw) > 2 else []

    blocks = []
    for k, hi in enumerate(header_indices):
        start = hi + 1
        end = header_indices[k + 1] if k + 1 < len(header_indices) else len(df_raw)
        if start >= end:
            continue
        header_row = df_raw.iloc[hi]
        block_df = df_raw.iloc[start:end].copy()
        num_cols = len(block_df.columns)
        col_names = []
        for j in range(num_cols):
            if j < len(header_row):
                v = header_row.iloc[j]
                col_names.append(str(v).strip() if pd.notna(v) else f"Unnamed_{j}")
            else:
                col_names.append(f"Unnamed_{j}")
        block_df.columns = col_names
        # 只對合併儲存格欄位做 ffill，Color code 等欄位保持原狀
        ffill_cols = _cols_to_ffill(col_names)
        if ffill_cols:
            block_df[ffill_cols] = block_df[ffill_cols].ffill(axis=0)
        blocks.append(block_df)
    return blocks


def _find_s_m_l_xl_columns(df: pd.DataFrame):
    """尋找 S, M, L, XL 欄位（依序），回傳 [col_S, col_M, col_L, col_XL] 或 None"""
    cols = []
    for name in ["S", "M", "L", "XL"]:
        found = None
        for c in df.columns:
            if str(c).strip() == name:
                found = c
                break
        if found is None:
            return None
        cols.append(found)
    return cols


def _safe_int(val, default: int = 0) -> int:
    if pd.isna(val):
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _get_size_quantity(row: pd.Series, df: pd.DataFrame, col_names: list, size_idx: int) -> int:
    """（保留供相容）取得指定尺寸的數量：大貨 + 樣品 + 5"""
    total = 0
    for col in col_names:
        if col in row.index and pd.notna(row.get(col)):
            try:
                total += int(float(row[col]))
            except (ValueError, TypeError):
                pass
    return total + 5


def _find_quantity_columns(df: pd.DataFrame) -> tuple:
    """（保留供相容）動態尋找 大貨交期 與 樣品交期 相關欄位"""
    headers = [str(c) for c in df.columns]
    dahuo = [c for c in headers if "大貨交期" in c]
    yangpin = [c for c in headers if "樣品交期" in c]
    if len(dahuo) >= 4 and len(yangpin) >= 4:
        return dahuo[:4], yangpin[:4]
    if len(dahuo) >= 4:
        return dahuo[:4], [yangpin[0]] * 4 if yangpin else [""] * 4
    if len(yangpin) >= 4:
        return [dahuo[0]] * 4 if dahuo else [""] * 4, yangpin[:4]
    d = dahuo[0] if dahuo else ""
    y = yangpin[0] if yangpin else ""
    return [d] * 4, [y] * 4


def _find_columns(df: pd.DataFrame) -> dict:
    """自動尋找所有需要的欄位，回傳欄位名稱 dict"""
    result = {}
    for c in df.columns:
        s = str(c).strip()
        sl = s.lower()
        if "color" in sl and "code" in sl:
            result["color_code"] = c
        if "style" in sl or s == "STYLE":
            result["style"] = c
        if "售價" in s or "價格" in s or "price" in sl:
            result["price"] = c
        if "顏色" in s and ("中" in s or "英" in s):
            result["color"] = c
        if "階段" in s:
            result["stage"] = c
    # 中文品名：優先完全匹配，排除顏色/英文
    for c in df.columns:
        s = str(c).strip()
        if "中文品名" in s:
            result["cn_name"] = c
            break
    if "cn_name" not in result:
        for c in df.columns:
            s = str(c).strip()
            if "品名" in s and "顏色" not in s and "英文" not in s:
                result["cn_name"] = c
                break
    # Fallback
    cols = df.columns
    result.setdefault("style", cols[1] if len(cols) > 1 else cols[0])
    result.setdefault("cn_name", cols[3] if len(cols) > 3 else cols[0])
    result.setdefault("price", cols[6] if len(cols) > 6 else cols[0])
    result.setdefault("color", cols[11] if len(cols) > 11 else cols[0])
    result.setdefault("color_code", cols[12] if len(cols) > 12 else cols[-1])
    return result


def process_sheet(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    處理單一資料區塊。
    用「階段」欄位篩選含有「交期」的列（排除加總列），
    依 (Style, Color code) 分組，加總所有大貨+樣品數量，每尺寸 +5。
    """
    col_map = _find_columns(df_raw)
    col_style = col_map["style"]
    col_cn = col_map["cn_name"]
    col_price = col_map["price"]
    col_color = col_map["color"]
    col_color_code = col_map["color_code"]
    col_stage = col_map.get("stage")

    size_cols = _find_s_m_l_xl_columns(df_raw)
    if not size_cols:
        return pd.DataFrame()

    # 篩選有效資料列：「階段」欄包含「交期」（排除加總列與空白列）
    if col_stage:
        mask = df_raw[col_stage].astype(str).str.contains("交期", na=False)
        df = df_raw[mask].copy()
    else:
        df = df_raw[
            df_raw[col_color_code].notna()
            & (df_raw[col_color_code].astype(str).str.strip() != "")
        ].copy()

    if df.empty:
        return pd.DataFrame()

    # 依 (STYLE 原始值, Color code) 分組，加總每尺寸數量
    rows_out = []
    seen_mpo = set()
    grouped = df.groupby([col_style, col_color_code], sort=False)
    for (style_raw, color_code), grp in grouped:
        style_id, mpo_id = _parse_style(style_raw)
        cc = str(color_code).strip()
        if not style_id or not cc:
            continue

        cn_name = grp[col_cn].iloc[0]
        price_val = _safe_int(grp[col_price].iloc[0])
        color_clean = _clean_color(str(grp[col_color].iloc[0]))

        # 檢查是否有大貨交期資料（排除「單純只有樣品交期」的顏色）
        if col_stage:
            has_dahuo = grp[col_stage].astype(str).str.contains("大貨", na=False).any()
            if not has_dahuo:
                continue

        # MPO#：同一個 MPO# 只在第一次出現時顯示
        show_mpo = mpo_id not in seen_mpo
        if mpo_id:
            seen_mpo.add(mpo_id)

        for idx, size in enumerate(SIZES):
            sc = size_cols[idx]
            total_qty = grp[sc].apply(lambda x: _safe_int(x)).sum() + 5
            style_number = f"{style_id}-{cc}-{size}"
            mpo_display = mpo_id if (size == "S" and show_mpo) else ""
            rows_out.append({
                "MPO#": mpo_display,
                "Style Number": style_number,
                "中文名稱": cn_name,
                "size": size,
                "COLOR": color_clean,
                "SKU": style_number,
                "Price (NT)": price_val,
                "吊牌貼紙數量(PCS)": total_qty,
                "單件袋貼紙數量(PCS)": total_qty,
            })

    return pd.DataFrame(rows_out)


def apply_formatting(ws, last_row: int) -> None:
    """套用輸出 Excel 的格式"""
    # 先建立「每列所屬 MPO block」與「block 顏色」對應
    current_mpo = None
    block_order = []
    row_to_block = {}
    for r in range(2, last_row + 1):
        a_val = ws.cell(r, 1).value
        if a_val and str(a_val).strip():
            current_mpo = str(a_val).strip()
            if current_mpo not in block_order:
                block_order.append(current_mpo)
        row_to_block[r] = current_mpo

    block_fill = {}
    for i, mpo in enumerate(block_order):
        block_fill[mpo] = BLOCK_FILL_1 if i % 2 == 0 else BLOCK_FILL_2

    # 標題列 Row 1：金色、置中、列高 25
    ws.row_dimensions[1].height = ROW1_HEIGHT
    for col in range(1, 12):
        c = ws.cell(1, col)
        c.font = Font(name=FONT_NAME, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")

    ws.merge_cells("H1:I1")
    ws.merge_cells("J1:K1")

    # 資料列：A 欄有 MPO# 才填橙色底，無資料則無底色
    for r in range(2, last_row + 1):
        mpo = row_to_block.get(r)
        fill = block_fill.get(mpo, BLOCK_FILL_1) if mpo else BLOCK_FILL_1
        a_cell = ws.cell(r, 1)
        has_mpo = a_cell.value and str(a_cell.value).strip()
        if has_mpo:
            a_cell.font = Font(name=FONT_NAME, size=11, color="FF0000")
            a_cell.fill = PatternFill(start_color=MPO_FILL, end_color=MPO_FILL, fill_type="solid")
        else:
            a_cell.font = Font(name=FONT_NAME, size=11)
            a_cell.fill = PatternFill()  # 無底色
        a_cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in range(2, 12):
            c = ws.cell(r, col)
            c.font = Font(name=FONT_NAME, size=11)
            c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            if col in (2, 3, 6):
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif col in (4, 5):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="right", vertical="center")

    # 欄寬 A=23, B=33, C=19, F=32 等
    for col, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _jpeg_to_png_bytes(data: bytes) -> bytes:
    """將 JPEG 轉成 PNG bytes。若已是 PNG 或無法辨識則回傳原 bytes。"""
    try:
        from PIL import Image
        img = Image.open(BytesIO(data))
        if img.format and img.format.upper() in ("JPEG", "JPG"):
            buf = BytesIO()
            img.convert("RGB").save(buf, format="PNG")
            buf.seek(0)
            return buf.getvalue()
    except Exception:
        pass
    return data


def _load_image_for_excel(
    source: Optional[Union[str, Path, bytes]],
    default_path: Path,
) -> Optional[OpenpyxlImage]:
    """從路徑或 bytes 載入圖片，供 Excel 使用。JPEG 會自動轉為 PNG。若皆無則嘗試預設路徑。"""
    if source is None:
        if default_path.exists():
            try:
                return OpenpyxlImage(str(default_path))
            except Exception:
                return None
        return None
    try:
        if isinstance(source, bytes):
            source = _jpeg_to_png_bytes(source)
            return OpenpyxlImage(BytesIO(source))
        path = Path(source) if isinstance(source, str) else source
        if path.exists():
            with open(path, "rb") as f:
                raw = f.read()
            raw = _jpeg_to_png_bytes(raw)
            return OpenpyxlImage(BytesIO(raw))
        return None
    except Exception:
        return None


def generate_output_excel(
    df: pd.DataFrame,
    hangtag_image: Optional[Union[str, Path, bytes]] = None,
    bag_image: Optional[Union[str, Path, bytes]] = None,
) -> bytes:
    """
    將處理後的 DataFrame 輸出為格式化的 Excel bytes。
    hangtag_image、bag_image 可為路徑或 bytes；若為 None 則使用專案內的預設檔。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "貼紙"

    ws.cell(1, 1, "MPO#")
    ws.cell(1, 2, "Style Number")
    ws.cell(1, 3, "中文名稱")
    ws.cell(1, 4, "size")
    ws.cell(1, 5, "COLOR")
    ws.cell(1, 6, "SKU")
    ws.cell(1, 7, "Price (NT)")
    ws.merge_cells("H1:I1")
    ws.cell(1, 8, "吊牌貼紙數量(PCS)")
    ws.merge_cells("J1:K1")
    ws.cell(1, 10, "單件袋貼紙數量(PCS)")

    # 資料
    for r, row in df.iterrows():
        excel_row = r + 2
        ws.cell(excel_row, 1, row["MPO#"])
        ws.cell(excel_row, 2, row["Style Number"])
        ws.cell(excel_row, 3, row["中文名稱"])
        ws.cell(excel_row, 4, row["size"])
        ws.cell(excel_row, 5, row["COLOR"])
        ws.cell(excel_row, 6, row["SKU"])
        ws.cell(excel_row, 7, row["Price (NT)"])
        ws.cell(excel_row, 8, row["吊牌貼紙數量(PCS)"])
        ws.cell(excel_row, 10, row["單件袋貼紙數量(PCS)"])

    last_row = len(df) + 1
    # 資料列 H&I、J&K 合併儲存格
    for r in range(2, last_row + 1):
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
        ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
    apply_formatting(ws, last_row)

    # 金額千位逗號
    for r in range(2, last_row + 1):
        ws.cell(r, 7).number_format = "#,##0"

    # 最後一列資料後隔一行空白列，插入吊牌貼紙與單件袋貼紙（固定實體尺寸）
    img_row = last_row + 2  # 空白列後開始
    hangtag_img = _load_image_for_excel(hangtag_image, _default_sticker_path("吊牌貼紙"))
    bag_img = _load_image_for_excel(bag_image, _default_sticker_path("單件袋貼紙"))

    col_c_half_px = (COLUMN_WIDTHS[2] / 2) * 7
    hangtag_offset_px = 25  # 吊牌貼紙往右偏移
    if hangtag_img:
        hangtag_anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=0,
                row=img_row - 1,
                colOff=pixels_to_EMU(hangtag_offset_px),
                rowOff=0,
            ),
            ext=XDRPositiveSize2D(
                cm_to_EMU(HANGTAG_WIDTH_CM),
                cm_to_EMU(HANGTAG_HEIGHT_CM),
            ),
        )
        ws.add_image(hangtag_img, hangtag_anchor)
    if bag_img:
        bag_anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=2,
                row=img_row - 1,
                colOff=pixels_to_EMU(col_c_half_px),
                rowOff=0,
            ),
            ext=XDRPositiveSize2D(
                cm_to_EMU(BAG_WIDTH_CM),
                cm_to_EMU(BAG_HEIGHT_CM),
            ),
        )
        ws.add_image(bag_img, bag_anchor)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def process_uploaded_excel(
    file_content: bytes,
    sheet_names: Optional[list[str]] = None,
    hangtag_image: Optional[Union[str, Path, bytes]] = None,
    bag_image: Optional[Union[str, Path, bytes]] = None,
) -> tuple[bytes, pd.DataFrame]:
    """
    處理上傳的 Excel 檔案。
    動態搜尋與 Row 3 相同的標題列，分段讀取資料後合併處理。
    hangtag_image、bag_image 可為路徑或 bytes；若為 None 則使用專案內的預設檔。
    """
    xl = pd.ExcelFile(BytesIO(file_content))
    all_sheets = xl.sheet_names

    if sheet_names:
        to_process = [s for s in sheet_names if s in all_sheets]
    else:
        to_process = all_sheets

    if not to_process:
        raise ValueError("沒有可處理的工作表")

    dfs = []
    for name in to_process:
        df_raw = pd.read_excel(xl, sheet_name=name, header=None)
        blocks = _get_blocks_from_sheet(df_raw)
        for block in blocks:
            if block.empty:
                continue
            df_out = process_sheet(block)
            if not df_out.empty:
                dfs.append(df_out)

    if not dfs:
        raise ValueError("沒有可輸出的資料，請確認工作表結構與標題列（需有 S, M, L, XL 欄位）")

    combined = pd.concat(dfs, ignore_index=True)
    excel_bytes = generate_output_excel(
        combined,
        hangtag_image=hangtag_image,
        bag_image=bag_image,
    )
    return excel_bytes, combined


def _build_source_expected(
    source_content: bytes,
    sheet_names: Optional[list[str]] = None,
) -> dict:
    """
    從來源 WIP 建立所有有效 Style+Color 的期望資料。
    回傳 dict，key = "StyleID-ColorCode"，value = {
        "cn_name": str, "price": int, "color": str,
        "qty": {"S": int, "M": int, "L": int, "XL": int},
    }
    """
    xl = pd.ExcelFile(BytesIO(source_content))
    to_process = (
        [s for s in sheet_names if s in xl.sheet_names]
        if sheet_names else xl.sheet_names
    )
    expected = {}
    for name in to_process:
        df_raw = pd.read_excel(xl, sheet_name=name, header=None)
        blocks = _get_blocks_from_sheet(df_raw)
        for block in blocks:
            if block.empty:
                continue
            col_map = _find_columns(block)
            col_style = col_map["style"]
            col_cn = col_map["cn_name"]
            col_price = col_map["price"]
            col_color = col_map["color"]
            col_color_code = col_map["color_code"]
            col_stage = col_map.get("stage")
            size_cols = _find_s_m_l_xl_columns(block)
            if not size_cols:
                continue

            if col_stage:
                mask = block[col_stage].astype(str).str.contains("交期", na=False)
                df = block[mask].copy()
            else:
                df = block[
                    block[col_color_code].notna()
                    & (block[col_color_code].astype(str).str.strip() != "")
                ].copy()
            if df.empty:
                continue

            grouped = df.groupby([col_style, col_color_code], sort=False)
            for (style_raw, cc), grp in grouped:
                if col_stage:
                    has_dahuo = grp[col_stage].astype(str).str.contains("大貨", na=False).any()
                    if not has_dahuo:
                        continue
                style_id, _ = _parse_style(style_raw)
                cc_str = str(cc).strip()
                if not style_id or not cc_str:
                    continue
                key = f"{style_id}-{cc_str}"
                cn_name = grp[col_cn].iloc[0]
                price_val = _safe_int(grp[col_price].iloc[0])
                color_clean = _clean_color(str(grp[col_color].iloc[0]))
                qty = {}
                for idx, size in enumerate(SIZES):
                    sc = size_cols[idx]
                    qty[size] = grp[sc].apply(lambda x: _safe_int(x)).sum() + 5
                expected[key] = {
                    "cn_name": cn_name,
                    "price": price_val,
                    "color": color_clean,
                    "qty": qty,
                }
    return expected


def verify_output(
    output_df: pd.DataFrame,
    source_content: bytes,
    sheet_names: Optional[list[str]] = None,
) -> list[dict]:
    """
    方案三：完整雙向比對。
    自動用同一份來源 WIP 驗證輸出，回傳驗證結果 list。
    每筆 dict: {"type": "pass"|"warn"|"error", "category": str, "message": str}
    """
    results = []

    if output_df.empty:
        results.append({"type": "error", "category": "結構", "message": "輸出資料為空"})
        return results

    # ── 1. 格式與結構檢查 ──

    # 1-a 每組 Style+Color 是否都有 S, M, L, XL
    output_df = output_df.copy()
    output_df["_style_color"] = output_df["Style Number"].apply(
        lambda x: x.rsplit("-", 1)[0]
    )
    for sc, grp in output_df.groupby("_style_color", sort=False):
        sizes_found = set(grp["size"].tolist())
        missing = set(SIZES) - sizes_found
        if missing:
            results.append({"type": "error", "category": "結構",
                            "message": f"{sc} 缺少尺寸：{', '.join(sorted(missing))}"})

    # 1-b COLOR 純中文
    for _, row in output_df.iterrows():
        color = str(row.get("COLOR", ""))
        if color and re.search(r"[a-zA-Z0-9#]", color):
            results.append({"type": "warn", "category": "格式",
                            "message": f"{row['Style Number']} 的 COLOR「{color}」含有非中文字元"})

    # 1-c Style Number 格式
    for _, row in output_df.iterrows():
        sn = str(row.get("Style Number", ""))
        if sn.count("-") < 2:
            results.append({"type": "warn", "category": "格式",
                            "message": f"Style Number「{sn}」格式異常（預期 [ID]-[Color]-[Size]）"})

    # 1-d 數量正整數
    for _, row in output_df.iterrows():
        for col in ["吊牌貼紙數量(PCS)", "單件袋貼紙數量(PCS)"]:
            v = row.get(col, 0)
            if not isinstance(v, (int, float)) or v < 0:
                results.append({"type": "warn", "category": "數量",
                                "message": f"{row['Style Number']} 的 {col} 為 {v}，異常"})

    # 1-e 售價有值
    for _, row in output_df.iterrows():
        if _safe_int(row.get("Price (NT)", 0)) <= 0:
            results.append({"type": "warn", "category": "格式",
                            "message": f"{row['Style Number']} 的售價為 0 或空"})

    struct_errors = [r for r in results if r["category"] == "結構" and r["type"] == "error"]
    if not struct_errors:
        results.append({"type": "pass", "category": "結構",
                        "message": "每組 Style+Color 皆有 S, M, L, XL 四筆"})
    format_warns = [r for r in results if r["category"] == "格式" and r["type"] != "pass"]
    if not format_warns:
        results.append({"type": "pass", "category": "格式", "message": "所有欄位格式正確"})

    # ── 2. 雙向比對（來源 vs 輸出）──
    try:
        expected = _build_source_expected(source_content, sheet_names)
    except Exception as e:
        results.append({"type": "warn", "category": "比對",
                        "message": f"無法讀取來源進行比對：{e}"})
        output_df.drop(columns=["_style_color"], inplace=True, errors="ignore")
        return results

    output_keys = set(output_df["_style_color"].unique())
    source_keys = set(expected.keys())

    # 2-a 來源有、輸出沒有 → 遺漏
    missing_in_output = source_keys - output_keys
    if missing_in_output:
        for k in sorted(missing_in_output):
            results.append({"type": "error", "category": "遺漏款式",
                            "message": f"{k} 存在於來源 WIP，但輸出中沒有"})
    else:
        results.append({"type": "pass", "category": "遺漏款式",
                        "message": "沒有遺漏，來源的每個款式+顏色都有輸出"})

    # 2-b 輸出有、來源沒有 → 多出
    extra_in_output = output_keys - source_keys
    if extra_in_output:
        for k in sorted(extra_in_output):
            results.append({"type": "error", "category": "多出款式",
                            "message": f"{k} 出現在輸出中，但來源 WIP 沒有此款式"})
    else:
        results.append({"type": "pass", "category": "多出款式",
                        "message": "沒有多出，輸出的每個款式+顏色都能對應來源"})

    # ── 3. 逐筆欄位值比對 ──
    qty_ok = True
    field_ok = True
    for _, row in output_df.iterrows():
        sc = row["_style_color"]
        exp = expected.get(sc)
        if exp is None:
            continue
        sn = row["Style Number"]
        size = row["size"]

        # 3-a 數量比對
        exp_qty = exp["qty"].get(size)
        if exp_qty is not None:
            actual_tag = _safe_int(row.get("吊牌貼紙數量(PCS)", 0))
            actual_bag = _safe_int(row.get("單件袋貼紙數量(PCS)", 0))
            if actual_tag != exp_qty:
                results.append({"type": "error", "category": "數量比對",
                                "message": f"{sn}：吊牌數量 預期={exp_qty}, 實際={actual_tag}"})
                qty_ok = False
            if actual_bag != exp_qty:
                results.append({"type": "error", "category": "數量比對",
                                "message": f"{sn}：單件袋數量 預期={exp_qty}, 實際={actual_bag}"})
                qty_ok = False

        # 3-b 中文品名
        out_cn = str(row.get("中文名稱", "")).strip()
        exp_cn = str(exp["cn_name"]).strip() if pd.notna(exp["cn_name"]) else ""
        if out_cn != exp_cn:
            results.append({"type": "error", "category": "欄位比對",
                            "message": f"{sn}：中文名稱 預期=「{exp_cn}」, 實際=「{out_cn}」"})
            field_ok = False

        # 3-c 售價
        out_price = _safe_int(row.get("Price (NT)", 0))
        exp_price = exp["price"]
        if out_price != exp_price:
            results.append({"type": "error", "category": "欄位比對",
                            "message": f"{sn}：售價 預期={exp_price}, 實際={out_price}"})
            field_ok = False

        # 3-d 顏色
        out_color = str(row.get("COLOR", "")).strip()
        exp_color = str(exp["color"]).strip()
        if out_color != exp_color:
            results.append({"type": "warn", "category": "欄位比對",
                            "message": f"{sn}：顏色 預期=「{exp_color}」, 實際=「{out_color}」"})
            field_ok = False

    if qty_ok:
        results.append({"type": "pass", "category": "數量比對",
                        "message": "所有數量與來源一致"})
    if field_ok:
        results.append({"type": "pass", "category": "欄位比對",
                        "message": "中文品名、售價、顏色皆與來源一致"})

    output_df.drop(columns=["_style_color"], inplace=True, errors="ignore")
    return results


def parse_output_filename_from_input(input_name: str) -> str:
    """
    從輸入檔名解析預設輸出檔名
    輸入範例: "CHOIR SS26 WIP - CW (T-Shirt) - 2026.01.21"
    輸出範例: "CHOIR SS26 - Sticler - CW - 2026.03.05"
    """
    from datetime import date
    today = date.today().strftime("%Y.%m.%d")
    s = (input_name or "").strip()
    if not s:
        return f"貼紙輸出 - {today}.xlsx"
    parts = [p.strip() for p in s.split("-")]
    if len(parts) >= 2:
        brand_season = parts[0].replace("WIP", "").strip()
        factory_part = parts[1]
        factory = factory_part.split("(")[0].strip() if "(" in factory_part else factory_part.strip()
        return f"{brand_season} - Sticler - {factory} - {today}.xlsx"
    return f"貼紙輸出 - {today}.xlsx"
